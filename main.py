import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import subprocess
import sys
import json
import shutil
from datetime import datetime, timedelta
import csv
import logging

# Настройка логирования (аудит)
LOG_DIR = os.path.join(os.path.expanduser('~'), 'Desktop', 'Учёт имущества')
os.makedirs(LOG_DIR, exist_ok=True)
logging.basicConfig(
    filename=os.path.join(LOG_DIR, 'audit.log'),
    level=logging.INFO,
    format='%(asctime)s - %(message)s'
)

class AssetManager:
    def __init__(self, root):
        self.root = root
        self.root.title("Учёт имущества")
        self.root.geometry("1400x800")
        self.root.configure(bg='#f0f0f0')

        self.dark_mode = False
        self.tooltip = None
        self.full_names = {}
        self.sort_column = None
        self.sort_reverse = False
        self.current_user = "user"   # без аутентификации

        # Пути
        desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
        self.data_dir = os.path.join(desktop, 'Учёт имущества')
        self.scans_dir = os.path.join(self.data_dir, 'Сканы')
        self.backup_dir = os.path.join(self.data_dir, 'Резервные копии')
        self.history_file = os.path.join(self.data_dir, 'history.json')
        self.last_backup_file = os.path.join(self.data_dir, 'last_backup.txt')
        for d in [self.scans_dir, self.backup_dir]:
            os.makedirs(d, exist_ok=True)
        self.excel_path = os.path.join(self.data_dir, 'assets.xlsx')
        self.settings_path = os.path.join(self.data_dir, 'settings.json')

        self.assets = []
        self.filtered_assets = []
        self.grouped_assets = []
        self.move_history = []

        # Загрузка настроек
        self.load_settings()

        # Загрузка истории перемещений
        self.load_move_history()

        if not os.path.exists(self.excel_path):
            self.create_excel_file()

        self.load_assets()
        self.setup_styles()
        self.setup_ui()
        self.apply_theme()
        self.setup_shortcuts()
        self.auto_save()
        self.schedule_backup()

        if hasattr(self, 'window_geometry'):
            self.root.geometry(self.window_geometry)

        # Первичное заполнение таблицы
        self.filter_assets()

    # ================== НАСТРОЙКИ ==================
    def load_settings(self):
        try:
            with open(self.settings_path, 'r', encoding='utf-8') as f:
                settings = json.load(f)
                self.dark_mode = settings.get('dark_mode', False)
                self.window_geometry = settings.get('window_geometry', '1400x800')
                self.column_widths = settings.get('column_widths', {})
        except:
            self.dark_mode = False
            self.window_geometry = '1400x800'
            self.column_widths = {}

    def save_settings(self):
        settings = {
            'dark_mode': self.dark_mode,
            'window_geometry': self.root.geometry(),
            'column_widths': self.get_column_widths()
        }
        try:
            with open(self.settings_path, 'w', encoding='utf-8') as f:
                json.dump(settings, f, ensure_ascii=False, indent=2)
        except:
            pass

    def get_column_widths(self):
        widths = {}
        for col in self.tree['columns']:
            widths[col] = self.tree.column(col, 'width')
        return widths

    def apply_saved_column_widths(self):
        if hasattr(self, 'column_widths'):
            for col, width in self.column_widths.items():
                if col in self.tree['columns']:
                    self.tree.column(col, width=width)

    # ================== ЦВЕТОВЫЕ СХЕМЫ ==================
    def get_colors(self):
        if self.dark_mode:
            return {
                'bg': '#2e2e2e', 'fg': '#ffffff', 'panel_bg': '#3c3c3c',
                'entry_bg': '#3c3c3c', 'entry_fg': '#ffffff',
                'button_bg': '#555555', 'button_fg': '#ffffff',
                'tree_bg': '#2e2e2e', 'tree_fg': '#ffffff',
                'tree_selected_bg': '#4a90d9', 'tree_selected_fg': '#ffffff',
                'heading_bg': '#1e1e1e', 'heading_fg': '#ffffff',
                'tag_written_off': '#5c3a3a', 'tag_in_repair': '#5c4a2a',
                'tag_lost': '#5c3a5c', 'tag_requires_repair': '#5c4a2a',
                'tag_reserve': '#4a4a4a', 'tag_in_use': '#2a5c2a',
                'filter_bg': '#3c3c3c', 'filter_fg': '#ffffff',
                'stats_bg': '#2e2e2e', 'stats_fg': '#cccccc',
                'detail_bg': '#2e2e2e', 'detail_fg': '#ffffff',
                'grid_color': '#555555'
            }
        else:
            return {
                'bg': '#f0f0f0', 'fg': '#333333', 'panel_bg': '#ffffff',
                'entry_bg': '#ffffff', 'entry_fg': '#333333',
                'button_bg': '#4CAF50', 'button_fg': '#ffffff',
                'tree_bg': '#ffffff', 'tree_fg': '#333333',
                'tree_selected_bg': '#2196F3', 'tree_selected_fg': '#ffffff',
                'heading_bg': '#2196F3', 'heading_fg': '#ffffff',
                'tag_written_off': '#FFEBEE', 'tag_in_repair': '#FFF3E0',
                'tag_lost': '#F3E5F5', 'tag_requires_repair': '#FFF3E0',
                'tag_reserve': '#F5F5F5', 'tag_in_use': '#E8F5E9',
                'filter_bg': '#ffffff', 'filter_fg': '#333333',
                'stats_bg': '#f0f0f0', 'stats_fg': '#666666',
                'detail_bg': '#f0f0f0', 'detail_fg': '#333333',
                'grid_color': '#CCCCCC'
            }

    def setup_styles(self):
        self.style = ttk.Style()
        self.style.theme_use('default')

    # ================== ПОСТРОЕНИЕ UI ==================
    def setup_ui(self):
        self.main_container = tk.Frame(self.root)
        self.main_container.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)

        # Верхняя панель
        self.top_panel = tk.Frame(self.main_container)
        self.top_panel.pack(fill=tk.X, pady=(0, 20))

        self.title_frame = tk.Frame(self.top_panel)
        self.title_frame.pack(side=tk.LEFT)

        self.title_label = tk.Label(self.title_frame, text="📦 Учёт имущества",
                                   font=('Segoe UI', 24, 'bold'))
        self.title_label.pack(side=tk.LEFT)

        self.subtitle_label = tk.Label(self.title_frame, text="by Ольгерд",
                                      font=('Segoe UI', 20))
        self.subtitle_label.pack(side=tk.LEFT, padx=(8, 0))

        self.button_frame = tk.Frame(self.top_panel)
        self.button_frame.pack(side=tk.RIGHT)

        self.add_btn = tk.Button(self.button_frame, text="➕ Добавить",
                                command=self.add_asset, relief=tk.FLAT,
                                cursor='hand2', padx=15, pady=8)
        self.add_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.edit_btn = tk.Button(self.button_frame, text="✏️ Редактировать",
                                 command=self.edit_asset, relief=tk.FLAT,
                                 cursor='hand2', padx=15, pady=8)
        self.edit_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.delete_btn = tk.Button(self.button_frame, text="🗑️ Удалить",
                                   command=self.delete_asset, relief=tk.FLAT,
                                   cursor='hand2', padx=15, pady=8)
        self.delete_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.move_btn = tk.Button(self.button_frame, text="🚚 Переместить",
                                 command=self.open_move_dialog, relief=tk.FLAT,
                                 cursor='hand2', padx=15, pady=8)
        self.move_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.history_btn = tk.Button(self.button_frame, text="📜 История",
                                    command=self.show_move_history, relief=tk.FLAT,
                                    cursor='hand2', padx=15, pady=8)
        self.history_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.inventory_btn = tk.Button(self.button_frame, text="📋 Инвентаризация",
                                      command=self.start_inventory, relief=tk.FLAT,
                                      cursor='hand2', padx=15, pady=8)
        self.inventory_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.import_btn = tk.Button(self.button_frame, text="📥 Импорт",
                                   command=self.import_from_file, relief=tk.FLAT,
                                   cursor='hand2', padx=15, pady=8)
        self.import_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.print_btn = tk.Button(self.button_frame, text="🖨️ Печать",
                                  command=self.print_report, relief=tk.FLAT,
                                  cursor='hand2', padx=15, pady=8)
        self.print_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.export_btn = tk.Button(self.button_frame, text="📊 Экспорт",
                                   command=self.export_to_excel, relief=tk.FLAT,
                                   cursor='hand2', padx=15, pady=8)
        self.export_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.help_btn = tk.Button(self.button_frame, text="❓",
                                 command=self.show_shortcuts_help, relief=tk.FLAT,
                                 cursor='hand2', padx=10, pady=8)
        self.help_btn.pack(side=tk.LEFT, padx=(5, 0))

        # Панель фильтров
        self.filter_panel = tk.Frame(self.main_container, relief=tk.RAISED, bd=1)
        self.filter_panel.pack(fill=tk.X, pady=(0, 20), padx=2)

        self.inner_filter = tk.Frame(self.filter_panel)
        self.inner_filter.pack(fill=tk.X, padx=10, pady=10)

        self.search_label = tk.Label(self.inner_filter, text="🔍 Поиск:")
        self.search_label.pack(side=tk.LEFT, padx=(0, 5))
        self.search_var = tk.StringVar()
        self.search_var.trace('w', lambda *args: self.filter_assets())
        self.search_entry = tk.Entry(self.inner_filter, textvariable=self.search_var, width=20)
        self.search_entry.pack(side=tk.LEFT, padx=(0, 10))

        self.name_label = tk.Label(self.inner_filter, text="Наименование:")
        self.name_label.pack(side=tk.LEFT, padx=(0, 5))
        self.name_filter_var = tk.StringVar()
        self.name_combo = ttk.Combobox(self.inner_filter, textvariable=self.name_filter_var,
                                       values=[""] + self.get_unique_values('name'),
                                       width=15, state='normal')
        self.name_combo.pack(side=tk.LEFT, padx=(0, 10))
        self.name_filter_var.trace('w', lambda *args: self.filter_assets())

        self.location_label = tk.Label(self.inner_filter, text="Расположение:")
        self.location_label.pack(side=tk.LEFT, padx=(0, 5))
        self.location_filter_var = tk.StringVar()
        self.location_combo = ttk.Combobox(self.inner_filter, textvariable=self.location_filter_var,
                                           values=[""] + self.get_unique_values('location'),
                                           width=15, state='normal')
        self.location_combo.pack(side=tk.LEFT, padx=(0, 10))
        self.location_filter_var.trace('w', lambda *args: self.filter_assets())

        self.act_label = tk.Label(self.inner_filter, text="Акт/Накладная:")
        self.act_label.pack(side=tk.LEFT, padx=(0, 5))
        self.act_filter_var = tk.StringVar()
        self.act_combo = ttk.Combobox(self.inner_filter, textvariable=self.act_filter_var,
                                      values=[""] + self.get_unique_values('act'),
                                      width=15, state='normal')
        self.act_combo.pack(side=tk.LEFT, padx=(0, 10))
        self.act_filter_var.trace('w', lambda *args: self.filter_assets())

        self.issued_to_label = tk.Label(self.inner_filter, text="Кому выдано:")
        self.issued_to_label.pack(side=tk.LEFT, padx=(0, 5))
        self.issued_to_filter_var = tk.StringVar()
        self.issued_to_combo = ttk.Combobox(self.inner_filter, textvariable=self.issued_to_filter_var,
                                            values=[""] + self.get_unique_values('issued_to'),
                                            width=15, state='normal')
        self.issued_to_combo.pack(side=tk.LEFT, padx=(0, 10))
        self.issued_to_filter_var.trace('w', lambda *args: self.filter_assets())

        self.direction_label = tk.Label(self.inner_filter, text="Направление:")
        self.direction_label.pack(side=tk.LEFT, padx=(0, 5))
        self.direction_var = tk.StringVar()
        self.direction_combo = ttk.Combobox(self.inner_filter, textvariable=self.direction_var,
                                            values=[""] + self.get_directions(),
                                            width=15, state='normal')
        self.direction_combo.pack(side=tk.LEFT, padx=(0, 10))
        self.direction_var.trace('w', lambda *args: self.filter_assets())

        self.status_label = tk.Label(self.inner_filter, text="Статус:")
        self.status_label.pack(side=tk.LEFT, padx=(0, 5))
        self.status_var = tk.StringVar()
        self.status_combo = ttk.Combobox(self.inner_filter, textvariable=self.status_var,
                                         values=["", "Активен", "В ремонте", "Списан", "Утеряно",
                                                 "Требует ремонта", "В резерве", "В эксплуатации"],
                                         width=12, state='normal')
        self.status_combo.pack(side=tk.LEFT)
        self.status_var.trace('w', lambda *args: self.filter_assets())

        # Таблица
        self.table_frame = tk.Frame(self.main_container, relief=tk.SOLID, bd=1)
        self.table_frame.pack(fill=tk.BOTH, expand=True)

        columns = {
            'name': ('Наименование', 350, 'w'),
            'inventory': ('Инв. номер', 100, 'center'),
            'qty_unit': ('Кол-во/ед. изм.', 100, 'center'),
            'direction': ('Направление', 130, 'w'),
            'location': ('Расположение', 150, 'w'),
            'issued_to': ('Кому выдано', 150, 'w'),
            'status': ('Статус', 100, 'center'),
            'act': ('Акт/Накладная', 130, 'w'),
            'commission_date': ('Дата ввода', 100, 'center'),
            'note': ('Примечание', 200, 'w')
        }

        self.tree = ttk.Treeview(self.table_frame, columns=tuple(columns.keys()),
                                 show='headings', selectmode='extended')

        for col, (text, width, anchor) in columns.items():
            self.tree.heading(col, text=text, command=lambda c=col: self.sort_by_column(c))
            self.tree.column(col, width=width, anchor=anchor, stretch=True)

        vscroll = ttk.Scrollbar(self.table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        hscroll = ttk.Scrollbar(self.table_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=vscroll.set, xscrollcommand=hscroll.set)

        self.tree.grid(row=0, column=0, sticky='nsew')
        vscroll.grid(row=0, column=1, sticky='ns')
        hscroll.grid(row=1, column=0, sticky='ew')
        self.table_frame.grid_rowconfigure(0, weight=1)
        self.table_frame.grid_columnconfigure(0, weight=1)

        self.tree.bind('<Double-Button-1>', self.on_tree_double_click)
        self.tree.bind('<Button-1>', self.on_tree_click)
        self.tree.bind('<<TreeviewSelect>>', self.on_tree_select)
        self.tree.bind('<Motion>', self.on_tree_motion)
        self.tree.bind('<Leave>', self.hide_tooltip)

        # Строка деталей
        self.detail_frame = tk.Frame(self.main_container)
        self.detail_frame.pack(fill=tk.X, pady=(5, 0))
        self.detail_label = tk.Label(self.detail_frame, text="", anchor='w',
                                     font=('Segoe UI', 10))
        self.detail_label.pack(fill=tk.X)

        # Статистика и кнопка темы
        self.stats_panel = tk.Frame(self.main_container)
        self.stats_panel.pack(fill=tk.X, pady=(10, 0))
        self.stats_label = tk.Label(self.stats_panel, text="")
        self.stats_label.pack(side=tk.LEFT)

        self.theme_btn = tk.Button(self.stats_panel, text="🌙 Тёмная тема",
                                   command=self.toggle_theme, relief=tk.FLAT,
                                   cursor='hand2', padx=10, pady=5)
        self.theme_btn.pack(side=tk.RIGHT)

        # Восстановление ширины столбцов
        self.apply_saved_column_widths()

    def apply_theme(self):
        c = self.get_colors()
        self.root.configure(bg=c['bg'])
        self.main_container.configure(bg=c['bg'])
        self.top_panel.configure(bg=c['bg'])
        self.title_frame.configure(bg=c['bg'])
        self.title_label.configure(bg=c['bg'], fg=c['fg'])
        self.subtitle_label.configure(bg=c['bg'], fg=c['stats_fg'])
        self.button_frame.configure(bg=c['bg'])
        self.filter_panel.configure(bg=c['filter_bg'])
        self.inner_filter.configure(bg=c['filter_bg'])
        self.table_frame.configure(bg=c['filter_bg'])
        self.stats_panel.configure(bg=c['bg'])
        self.stats_label.configure(bg=c['bg'], fg=c['stats_fg'])
        self.detail_frame.configure(bg=c['detail_bg'])
        self.detail_label.configure(bg=c['detail_bg'], fg=c['detail_fg'])

        for btn in [self.add_btn, self.edit_btn, self.delete_btn, self.move_btn,
                    self.history_btn, self.inventory_btn, self.import_btn,
                    self.print_btn, self.export_btn, self.help_btn, self.theme_btn]:
            btn.configure(bg=c['button_bg'], fg=c['button_fg'],
                         activebackground=c['button_bg'])

        for label in [self.search_label, self.name_label, self.location_label,
                      self.act_label, self.issued_to_label, self.direction_label, self.status_label]:
            label.configure(bg=c['filter_bg'], fg=c['filter_fg'])

        self.search_entry.configure(bg=c['entry_bg'], fg=c['entry_fg'],
                                    insertbackground=c['entry_fg'])

        self.style.configure('Treeview',
                             background=c['tree_bg'],
                             fieldbackground=c['tree_bg'],
                             foreground=c['tree_fg'],
                             rowheight=40,
                             font=('Segoe UI', 10),
                             borderwidth=1,
                             relief='solid',
                             bordercolor=c['grid_color'],
                             lightcolor=c['grid_color'],
                             darkcolor=c['grid_color'])
        self.style.configure('Treeview.Heading',
                             background=c['heading_bg'],
                             foreground=c['heading_fg'],
                             font=('Segoe UI', 10, 'bold'),
                             relief='solid',
                             borderwidth=1,
                             bordercolor=c['grid_color'],
                             lightcolor=c['grid_color'],
                             darkcolor=c['grid_color'])
        self.style.map('Treeview',
                       background=[('selected', c['tree_selected_bg'])],
                       foreground=[('selected', c['tree_selected_fg'])])
        self.style.map('Treeview.Heading',
                       background=[('active', c['heading_bg'])])
        self.style.configure('TCombobox',
                             fieldbackground=c['entry_bg'],
                             background=c['entry_bg'],
                             foreground=c['entry_fg'],
                             arrowcolor=c['entry_fg'])

        self.tree.tag_configure('written_off', background=c['tag_written_off'])
        self.tree.tag_configure('in_repair', background=c['tag_in_repair'])
        self.tree.tag_configure('lost', background=c['tag_lost'])
        self.tree.tag_configure('requires_repair', background=c['tag_requires_repair'])
        self.tree.tag_configure('reserve', background=c['tag_reserve'])
        self.tree.tag_configure('in_use', background=c['tag_in_use'])
        self.tree.tag_configure('group', font=('Segoe UI', 10, 'bold'))

        self.theme_btn.config(text="☀️ Светлая тема" if self.dark_mode else "🌙 Тёмная тема")

    def toggle_theme(self):
        self.dark_mode = not self.dark_mode
        self.apply_theme()

    # ================== TOOLTIP ==================
    def on_tree_motion(self, event):
        region = self.tree.identify('region', event.x, event.y)
        if region == 'cell':
            col = self.tree.identify_column(event.x)
            row = self.tree.identify_row(event.y)
            if row and col:
                col_index = int(col.replace('#', '')) - 1
                if col_index == 0:
                    full_name = self.full_names.get(row, '')
                    if full_name and len(full_name) > 20:
                        self.show_tooltip(event.x_root, event.y_root, full_name)
                    else:
                        self.hide_tooltip()
                else:
                    values = self.tree.item(row, 'values')
                    if 0 <= col_index < len(values):
                        text = str(values[col_index])
                        if len(text) > 20:
                            self.show_tooltip(event.x_root, event.y_root, text)
                        else:
                            self.hide_tooltip()
                    else:
                        self.hide_tooltip()
            else:
                self.hide_tooltip()
        else:
            self.hide_tooltip()

    def show_tooltip(self, x, y, text):
        self.hide_tooltip()
        self.tooltip = tk.Toplevel(self.root)
        self.tooltip.wm_overrideredirect(True)
        self.tooltip.wm_geometry(f"+{x+10}+{y+10}")
        label = tk.Label(self.tooltip, text=text, background='#ffffe0',
                         relief='solid', borderwidth=1, padx=5, pady=2,
                         wraplength=400, justify='left', fg='#000000')
        label.pack()

    def hide_tooltip(self, event=None):
        if self.tooltip:
            self.tooltip.destroy()
            self.tooltip = None

    # ================== СТРОКА ДЕТАЛЕЙ ==================
    def on_tree_select(self, event):
        selected = self.tree.selection()
        if not selected:
            self.detail_label.config(text="")
            return
        item = selected[0]
        values = self.tree.item(item, 'values')
        children = self.tree.get_children(item)
        if children:
            name = self.full_names.get(item, '')
            qty = values[2] if len(values) > 2 else ""
            self.detail_label.config(text=f"Группа: {name} | Всего: {qty}")
        else:
            name = self.full_names.get(item, '')
            inv = values[1] if len(values) > 1 else ""
            self.detail_label.config(text=f"Актив: {name} | Инв. номер: {inv}")

    # ================== ОБРЕЗКА ТЕКСТА ==================
    def truncate_text(self, text, max_length=30):
        if len(text) > max_length:
            return text[:max_length-1] + '…'
        return text

    # ================== ВСПОМОГАТЕЛЬНЫЕ МЕТОДЫ ==================
    def get_unique_values(self, field):
        values = set()
        for asset in self.assets:
            val = asset.get(field, '')
            if val:
                values.add(str(val))
        return sorted(list(values))

    def get_directions(self):
        return self.get_unique_values('direction')

    def update_filter_values(self):
        self.name_combo['values'] = [""] + self.get_unique_values('name')
        self.location_combo['values'] = [""] + self.get_unique_values('location')
        self.act_combo['values'] = [""] + self.get_unique_values('act')
        self.issued_to_combo['values'] = [""] + self.get_unique_values('issued_to')
        self.direction_combo['values'] = [""] + self.get_directions()

    # ================== РАБОТА С EXCEL ==================
    def create_excel_file(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Активы"
        headers = [
            'Наименование', 'Инв. номер', 'Количество', 'Ед. измерения',
            'Направление', 'Расположение', 'Кому выдано', 'Статус',
            'Акт/Накладная', 'Дата ввода', 'Примечание'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF', size=12)
            cell.fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        widths = [25, 15, 10, 12, 20, 20, 20, 12, 20, 15, 25]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        wb.save(self.excel_path)

    def load_assets(self):
        try:
            wb = openpyxl.load_workbook(self.excel_path)
            ws = wb.active
            self.assets = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[0]:
                    asset = {
                        'name': row[0] or '',
                        'inventory': row[1] or '',
                        'quantity': row[2] if isinstance(row[2], (int, float)) else 1,
                        'unit': row[3] or 'шт.',
                        'direction': row[4] or '',
                        'location': row[5] or '',
                        'issued_to': row[6] or '',
                        'status': row[7] or 'Активен',
                        'act': row[8] or '',
                        'commission_date': row[9] if len(row) > 9 else '',
                        'note': row[10] if len(row) > 10 else ''
                    }
                    self.assets.append(asset)
            self.filtered_assets = self.assets.copy()
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось загрузить данные: {str(e)}")

    def save_to_excel(self):
        try:
            self.backup_excel()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Активы"
            headers = [
                'Наименование', 'Инв. номер', 'Количество', 'Ед. измерения',
                'Направление', 'Расположение', 'Кому выдано', 'Статус',
                'Акт/Накладная', 'Дата ввода', 'Примечание'
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF', size=12)
                cell.fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row_idx, asset in enumerate(self.assets, 2):
                ws.cell(row=row_idx, column=1, value=asset['name'])
                ws.cell(row=row_idx, column=2, value=asset['inventory'])
                ws.cell(row=row_idx, column=3, value=asset['quantity'])
                ws.cell(row=row_idx, column=4, value=asset['unit'])
                ws.cell(row=row_idx, column=5, value=asset['direction'])
                ws.cell(row=row_idx, column=6, value=asset['location'])
                ws.cell(row=row_idx, column=7, value=asset['issued_to'])
                ws.cell(row=row_idx, column=8, value=asset['status'])
                ws.cell(row=row_idx, column=9, value=asset['act'])
                ws.cell(row=row_idx, column=10, value=asset['commission_date'])
                ws.cell(row=row_idx, column=11, value=asset['note'])

                if row_idx % 2 == 0:
                    for col in range(1, 12):
                        ws.cell(row=row_idx, column=col).fill = PatternFill(
                            start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

            widths = [25, 15, 10, 12, 20, 20, 20, 12, 20, 15, 25]
            for col, width in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width

            ws.auto_filter.ref = f"A1:K{len(self.assets) + 1}"
            ws.freeze_panes = 'A2'
            wb.save(self.excel_path)
            return True
        except PermissionError:
            backup_path = os.path.join(os.path.expanduser('~'), 'AppData', 'Local', 'Temp', 'assets_backup.xlsx')
            try:
                wb.save(backup_path)
                messagebox.showwarning(
                    "Внимание",
                    f"Файл {self.excel_path} занят другим процессом (возможно, открыт в Excel).\n"
                    f"Данные сохранены во временный файл:\n{backup_path}\n\n"
                    "Закройте Excel и повторите сохранение, либо скопируйте временный файл вручную."
                )
            except Exception as e2:
                messagebox.showerror("Ошибка", f"Не удалось сохранить данные: {str(e2)}")
            return False
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить данные: {str(e)}")
            return False

    # ================== РЕЗЕРВНОЕ КОПИРОВАНИЕ ==================
    def backup_excel(self):
        if not os.path.exists(self.excel_path):
            return
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"backup_{timestamp}.xlsx"
        backup_path = os.path.join(self.backup_dir, backup_name)
        shutil.copy2(self.excel_path, backup_path)
        backups = sorted([f for f in os.listdir(self.backup_dir) if f.startswith("backup_")])
        while len(backups) > 5:
            old = backups.pop(0)
            try:
                os.remove(os.path.join(self.backup_dir, old))
            except:
                pass

    def schedule_backup(self):
        if os.path.exists(self.last_backup_file):
            try:
                with open(self.last_backup_file, 'r') as f:
                    last_str = f.read().strip()
                    # Преобразуем строку в datetime, затем берём только дату
                    last_date = datetime.strptime(last_str, "%Y-%m-%d").date()
            except Exception:
                last_date = None
        else:
            last_date = None

        today = datetime.now().date()
        if last_date is None or last_date < today:
            self.backup_excel()
            with open(self.last_backup_file, 'w') as f:
                f.write(today.strftime("%Y-%m-%d"))

    # ================== ГРУППИРОВКА ==================
    def build_grouped_assets(self):
        groups = {}
        for asset in self.filtered_assets:
            key = asset['name'].strip().lower()
            if key not in groups:
                groups[key] = {
                    'name': asset['name'],
                    'assets': [],
                    'total_qty': 0,
                    'unit': asset['unit'],
                    'inventories': []
                }
            groups[key]['assets'].append(asset)
            groups[key]['total_qty'] += asset['quantity']
            groups[key]['inventories'].append(asset['inventory'])
        self.grouped_assets = list(groups.values())

    # ================== ОБРАБОТЧИКИ СОБЫТИЙ ==================
    def on_tree_click(self, event):
        region = self.tree.identify('region', event.x, event.y)
        if region == 'cell':
            col = self.tree.identify_column(event.x)
            if col == '#8':
                item = self.tree.identify_row(event.y)
                if item:
                    values = self.tree.item(item, 'values')
                    if values and len(values) >= 8:
                        act = values[7]
                        self.open_scan(act)

    def on_tree_double_click(self, event):
        item = self.tree.identify_row(event.y)
        if not item:
            return
        children = self.tree.get_children(item)
        if children:
            current_state = self.tree.item(item, 'open')
            self.tree.item(item, open=not current_state)
            values = list(self.tree.item(item, 'values'))
            if values:
                full_name = self.full_names.get(item, '')
                if not current_state:
                    values[0] = self.truncate_text('▾ ' + full_name)
                else:
                    values[0] = self.truncate_text('▸ ' + full_name)
                self.tree.item(item, values=values)
            return "break"
        else:
            self.edit_asset_by_item(item)
            return "break"

    def open_scan(self, act_name):
        if not act_name:
            return
        target = None
        for f in os.listdir(self.scans_dir):
            if act_name.lower() in f.lower():
                target = os.path.join(self.scans_dir, f)
                break
        if target and os.path.exists(target):
            try:
                if sys.platform == 'win32':
                    os.startfile(target)
                elif sys.platform == 'darwin':
                    subprocess.call(['open', target])
                else:
                    subprocess.call(['xdg-open', target])
            except Exception as e:
                messagebox.showerror("Ошибка", f"Не удалось открыть файл: {str(e)}")
        else:
            messagebox.showinfo("Информация", f"Скан не найден. Разместите файл в папке:\n{self.scans_dir}\n\nИскомое название: {act_name}")

    def has_scan(self, act_name):
        if not act_name:
            return False
        for f in os.listdir(self.scans_dir):
            if act_name.lower() in f.lower():
                return True
        return False

    # ================== CRUD ==================
    def add_asset(self):
        dialog = AssetDialog(self.root, "Добавить актив")
        self.root.wait_window(dialog.dialog)
        if dialog.result:
            if not dialog.result.get('inventory'):
                dialog.result['inventory'] = self.generate_inventory_number()
            self.assets.append(dialog.result)
            self.log_action("add", f"Добавлен актив {dialog.result['name']} ({dialog.result['inventory']})")
            self.save_to_excel()
            self.refresh_after_change()
            messagebox.showinfo("Успех", "Актив успешно добавлен")

    def edit_asset(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Внимание", "Выберите актив для редактирования")
            return
        item = selected[0]
        self.edit_asset_by_item(item)

    def edit_asset_by_item(self, item):
        children = self.tree.get_children(item)
        if children:
            messagebox.showinfo("Информация", "Выберите конкретный актив внутри группы")
            return
        values = self.tree.item(item, 'values')
        if not values or len(values) < 2:
            messagebox.showerror("Ошибка", "Не удалось определить актив")
            return
        inventory = values[1]
        asset_index = None
        for i, asset in enumerate(self.assets):
            if str(asset['inventory']) == str(inventory):
                asset_index = i
                break
        if asset_index is None:
            messagebox.showerror("Ошибка", "Актив не найден в базе данных")
            return
        dialog = AssetDialog(self.root, "Редактировать актив", self.assets[asset_index])
        self.root.wait_window(dialog.dialog)
        if dialog.result:
            self.log_action("edit", f"Изменён актив {self.assets[asset_index]['name']} ({inventory})")
            self.assets[asset_index] = dialog.result
            self.save_to_excel()
            self.refresh_after_change()
            messagebox.showinfo("Успех", "Актив успешно обновлён")

    def delete_asset(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Внимание", "Выберите активы для удаления")
            return
        if not messagebox.askyesno("Подтверждение",
                                   f"Удалить выбранные активы ({len(selected)} шт.)?"):
            return
        inv_numbers = []
        for item in selected:
            children = self.tree.get_children(item)
            if children:
                for child in children:
                    vals = self.tree.item(child, 'values')
                    if vals and len(vals) >= 2:
                        inv_numbers.append(vals[1])
            else:
                vals = self.tree.item(item, 'values')
                if vals and len(vals) >= 2:
                    inv_numbers.append(vals[1])
        if not inv_numbers:
            return
        self.log_action("delete", f"Удалены активы: {', '.join(inv_numbers)}")
        self.assets = [a for a in self.assets if str(a['inventory']) not in inv_numbers]
        self.save_to_excel()
        self.refresh_after_change()
        messagebox.showinfo("Успех", "Активы успешно удалены")

    def generate_inventory_number(self):
        if not self.assets:
            return "INV-0001"
        max_num = 0
        for asset in self.assets:
            inv = str(asset['inventory'])
            if inv.startswith("INV-"):
                try:
                    num = int(inv.split("-")[1])
                    max_num = max(max_num, num)
                except:
                    pass
        return f"INV-{max_num + 1:04d}"

    # ================== ФИЛЬТРАЦИЯ И ОТОБРАЖЕНИЕ ==================
    def refresh_after_change(self):
        self.update_filter_values()
        self.filter_assets()

    def filter_assets(self):
        search_text = self.search_var.get().strip().lower()
        name_filter = self.name_filter_var.get().strip().lower()
        location_filter = self.location_filter_var.get().strip().lower()
        act_filter = self.act_filter_var.get().strip().lower()
        issued_to_filter = self.issued_to_filter_var.get().strip().lower()
        direction_filter = self.direction_var.get().strip().lower()
        status_filter = self.status_var.get().strip()

        self.filtered_assets = []
        for asset in self.assets:
            if name_filter and asset['name'].lower() != name_filter:
                continue
            if location_filter and asset['location'].lower() != location_filter:
                continue
            if act_filter and asset['act'].lower() != act_filter:
                continue
            if issued_to_filter and asset['issued_to'].lower() != issued_to_filter:
                continue
            if direction_filter and asset['direction'].lower() != direction_filter:
                continue
            if status_filter and asset['status'] != status_filter:
                continue
            if search_text:
                searchable = ' '.join([
                    str(asset.get('name', '')),
                    str(asset.get('inventory', '')),
                    str(asset.get('location', '')),
                    str(asset.get('direction', '')),
                    str(asset.get('act', '')),
                    str(asset.get('issued_to', '')),
                    str(asset.get('note', ''))
                ]).lower()
                if search_text not in searchable:
                    continue
            self.filtered_assets.append(asset)

        # Применяем сортировку, если она задана
        if self.sort_column:
            key_map = {
                'name': 'name',
                'inventory': 'inventory',
                'qty_unit': 'quantity',
                'direction': 'direction',
                'location': 'location',
                'issued_to': 'issued_to',
                'status': 'status',
                'act': 'act',
                'commission_date': 'commission_date',
                'note': 'note'
            }
            key = key_map.get(self.sort_column, 'name')
            self.filtered_assets.sort(key=lambda a: str(a.get(key, '')), reverse=self.sort_reverse)

        self.build_grouped_assets()
        self.update_table()
        self.update_stats()

    def update_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.full_names.clear()

        for group in self.grouped_assets:
            if len(group['assets']) == 1:
                asset = group['assets'][0]
                tags = self.get_tags(asset)
                display_name = self.truncate_text(asset['name'])
                act_display = asset.get('act', '')
                if self.has_scan(act_display):
                    act_display = '📎 ' + act_display
                item_id = self.tree.insert('', 'end', values=(
                    display_name,
                    asset['inventory'],
                    f"{asset['quantity']}/{asset['unit']}",
                    asset['direction'],
                    asset['location'],
                    asset['issued_to'],
                    asset['status'],
                    act_display,
                    asset.get('commission_date', ''),
                    asset['note']
                ), tags=tags)
                self.full_names[item_id] = asset['name']
            else:
                group_tags = ('group',)
                group_display_name = self.truncate_text('▸ ' + group['name'])
                group_id = self.tree.insert('', 'end', values=(
                    group_display_name,
                    '',
                    f"{group['total_qty']}/{group['unit']}",
                    '',
                    '',
                    '',
                    '',
                    '',
                    '',
                    ''
                ), tags=group_tags, open=False)   # свёрнуто по умолчанию
                self.full_names[group_id] = group['name']

                for asset in group['assets']:
                    tags = self.get_tags(asset)
                    act_display = asset.get('act', '')
                    if self.has_scan(act_display):
                        act_display = '📎 ' + act_display
                    child_id = self.tree.insert(group_id, 'end', values=(
                        '',
                        asset['inventory'],
                        f"{asset['quantity']}/{asset['unit']}",
                        asset['direction'],
                        asset['location'],
                        asset['issued_to'],
                        asset['status'],
                        act_display,
                        asset.get('commission_date', ''),
                        asset['note']
                    ), tags=tags)
                    self.full_names[child_id] = asset['name']

    def get_tags(self, asset):
        tags = []
        status = asset.get('status', '')
        if status == 'Списан':
            tags.append('written_off')
        elif status == 'В ремонте':
            tags.append('in_repair')
        elif status == 'Утеряно':
            tags.append('lost')
        elif status == 'Требует ремонта':
            tags.append('requires_repair')
        elif status == 'В резерве':
            tags.append('reserve')
        elif status == 'В эксплуатации':
            tags.append('in_use')
        return tags

    def update_stats(self):
        total = len(self.assets)
        active = len([a for a in self.assets if a['status'] == 'Активен'])
        in_repair = len([a for a in self.assets if a['status'] == 'В ремонте'])
        written_off = len([a for a in self.assets if a['status'] == 'Списан'])
        lost = len([a for a in self.assets if a['status'] == 'Утеряно'])
        self.stats_label.config(
            text=f"Всего: {total} | Активных: {active} | В ремонте: {in_repair} | "
                 f"Списано: {written_off} | Утеряно: {lost}"
        )

    # ================== СОРТИРОВКА ==================
    def sort_by_column(self, col):
        if self.sort_column == col:
            self.sort_reverse = not self.sort_reverse
        else:
            self.sort_column = col
            self.sort_reverse = False
        self.filter_assets()   # пересоберёт с учётом сортировки

    # ================== ИСТОРИЯ ПЕРЕМЕЩЕНИЙ ==================
    def load_move_history(self):
        if os.path.exists(self.history_file):
            try:
                with open(self.history_file, 'r', encoding='utf-8') as f:
                    self.move_history = json.load(f)
            except:
                self.move_history = []
        else:
            self.move_history = []

    def save_move_history(self):
        try:
            with open(self.history_file, 'w', encoding='utf-8') as f:
                json.dump(self.move_history, f, ensure_ascii=False, indent=2)
        except:
            pass

    def record_move(self, asset_inv, from_loc, to_loc, date, user, note=""):
        record = {
            "asset_inv": asset_inv,
            "from": from_loc,
            "to": to_loc,
            "date": date,
            "user": user,
            "note": note
        }
        self.move_history.append(record)
        self.save_move_history()
        self.log_action("move", f"{asset_inv}: {from_loc} -> {to_loc}")

    def open_move_dialog(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("Внимание", "Выберите актив для перемещения")
            return
        item = selected[0]
        values = self.tree.item(item, 'values')
        if not values or len(values) < 2:
            messagebox.showerror("Ошибка", "Не удалось определить актив")
            return
        inv = values[1]
        asset = None
        for a in self.assets:
            if str(a['inventory']) == str(inv):
                asset = a
                break
        if not asset:
            messagebox.showerror("Ошибка", "Актив не найден")
            return
        dialog = MoveDialog(self.root, asset, self)
        self.root.wait_window(dialog.dialog)

    def show_move_history(self):
        if not self.move_history:
            messagebox.showinfo("История", "История перемещений пуста")
            return
        win = tk.Toplevel(self.root)
        win.title("История перемещений")
        win.geometry("800x400")
        text = tk.Text(win)
        text.pack(fill=tk.BOTH, expand=True)
        for rec in self.move_history:
            text.insert(tk.END, f"{rec['date']} | {rec['asset_inv']} | {rec['from']} -> {rec['to']} | {rec['user']} | {rec['note']}\n")
        text.config(state=tk.DISABLED)

    # ================== ИНВЕНТАРИЗАЦИЯ ==================
    def start_inventory(self):
        locations = self.get_unique_values('location')
        choice = simpledialog.askstring("Инвентаризация", "Введите расположение (или оставьте пустым для всех):")
        if choice is None:
            return
        if choice == "":
            inv_assets = self.assets
        else:
            inv_assets = [a for a in self.assets if a['location'] == choice]

        if not inv_assets:
            messagebox.showinfo("Инвентаризация", "Нет активов для выбранного расположения")
            return

        inv_dialog = tk.Toplevel(self.root)
        inv_dialog.title("Инвентаризация")
        inv_dialog.geometry("600x500")
        inv_dialog.grab_set()

        result = {}
        canvas = tk.Canvas(inv_dialog)
        scrollbar = ttk.Scrollbar(inv_dialog, orient="vertical", command=canvas.yview)
        frame = tk.Frame(canvas)
        canvas.create_window((0, 0), window=frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)

        for asset in inv_assets:
            var = tk.BooleanVar(value=True)
            result[asset['inventory']] = var
            row = tk.Frame(frame)
            row.pack(fill=tk.X, pady=2)
            tk.Checkbutton(row, variable=var).pack(side=tk.LEFT)
            tk.Label(row, text=f"{asset['name']} ({asset['inventory']}) - {asset['location']}").pack(side=tk.LEFT)

        def finish():
            discrepancies = []
            for asset in inv_assets:
                if not result[asset['inventory']].get():
                    discrepancies.append({
                        'inventory': asset['inventory'],
                        'name': asset['name'],
                        'location': asset['location'],
                        'status': 'Недостача'
                    })
            if discrepancies:
                file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                         filetypes=[("Excel files", "*.xlsx")])
                if file_path:
                    self.create_discrepancy_report(discrepancies, file_path)
            messagebox.showinfo("Готово", "Инвентаризация завершена")
            inv_dialog.destroy()

        tk.Button(inv_dialog, text="Завершить", command=finish).pack(pady=10)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        frame.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))

    def create_discrepancy_report(self, discrepancies, file_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Сличительная ведомость"
        headers = ['Инв. номер', 'Наименование', 'Расположение', 'Причина']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True)
        for row, disc in enumerate(discrepancies, 2):
            ws.cell(row=row, column=1, value=disc['inventory'])
            ws.cell(row=row, column=2, value=disc['name'])
            ws.cell(row=row, column=3, value=disc['location'])
            ws.cell(row=row, column=4, value=disc['status'])
        wb.save(file_path)

    # ================== ИМПОРТ ==================
    def import_from_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Excel/CSV files", "*.xlsx *.csv")])
        if not file_path:
            return
        try:
            if file_path.lower().endswith('.csv'):
                self.import_from_csv(file_path)
            else:
                self.import_from_excel(file_path)
            self.save_to_excel()
            self.refresh_after_change()
            messagebox.showinfo("Успех", "Импорт завершён")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось импортировать: {str(e)}")

    def import_from_excel(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                asset = {
                    'name': row[0],
                    'inventory': row[1] or self.generate_inventory_number(),
                    'quantity': row[2] if isinstance(row[2], (int, float)) else 1,
                    'unit': row[3] or 'шт.',
                    'direction': row[4] or '',
                    'location': row[5] or '',
                    'issued_to': row[6] or '',
                    'status': row[7] or 'Активен',
                    'act': row[8] or '',
                    'commission_date': row[9] if len(row) > 9 else '',
                    'note': row[10] if len(row) > 10 else ''
                }
                self.assets.append(asset)

    def import_from_csv(self, file_path):
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader, None)
            for row in reader:
                if row:
                    asset = {
                        'name': row[0],
                        'inventory': row[1] or self.generate_inventory_number(),
                        'quantity': int(row[2]) if row[2] else 1,
                        'unit': row[3] or 'шт.',
                        'direction': row[4] or '',
                        'location': row[5] or '',
                        'issued_to': row[6] or '',
                        'status': row[7] or 'Активен',
                        'act': row[8] or '',
                        'commission_date': row[9] if len(row) > 9 else '',
                        'note': row[10] if len(row) > 10 else ''
                    }
                    self.assets.append(asset)

    # ================== ПЕЧАТЬ ОТЧЁТА ==================
    def print_report(self):
        try:
            from reportlab.lib.pagesizes import A4
            from reportlab.lib import colors
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
            from reportlab.lib.styles import getSampleStyleSheet
        except ImportError:
            messagebox.showerror("Ошибка", "Для печати необходимо установить reportlab\npip install reportlab")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".pdf",
                                                 filetypes=[("PDF files", "*.pdf")])
        if not file_path:
            return

        doc = SimpleDocTemplate(file_path, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph("Отчёт по активам", styles['Title']))
        elements.append(Paragraph(f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}", styles['Normal']))
        elements.append(Paragraph("", styles['Normal']))

        data = [['Наименование', 'Инв. номер', 'Кол-во', 'Ед.', 'Направление', 'Расположение', 'Кому выдано', 'Статус']]
        for asset in self.filtered_assets:
            data.append([
                asset['name'],
                asset['inventory'],
                str(asset['quantity']),
                asset['unit'],
                asset['direction'],
                asset['location'],
                asset['issued_to'],
                asset['status']
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(table)
        doc.build(elements)
        messagebox.showinfo("Успех", f"Отчёт сохранён в {file_path}")
        if sys.platform == 'win32':
            os.startfile(file_path)

    # ================== АУДИТ ==================
    def log_action(self, action, details):
        logging.info(f"{action} - {details}")

    # ================== ЭКСПОРТ ==================
    def export_to_excel(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="filtered_export.xlsx"
        )
        if not file_path:
            return
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Отфильтрованные активы"
            headers = [
                'Наименование', 'Инв. номер', 'Количество', 'Ед. измерения',
                'Направление', 'Расположение', 'Кому выдано', 'Статус',
                'Акт/Накладная', 'Дата ввода', 'Примечание'
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color='FFFFFF', size=12)
                cell.fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            for row_idx, asset in enumerate(self.filtered_assets, 2):
                ws.cell(row=row_idx, column=1, value=asset['name'])
                ws.cell(row=row_idx, column=2, value=asset['inventory'])
                ws.cell(row=row_idx, column=3, value=asset['quantity'])
                ws.cell(row=row_idx, column=4, value=asset['unit'])
                ws.cell(row=row_idx, column=5, value=asset['direction'])
                ws.cell(row=row_idx, column=6, value=asset['location'])
                ws.cell(row=row_idx, column=7, value=asset['issued_to'])
                ws.cell(row=row_idx, column=8, value=asset['status'])
                ws.cell(row=row_idx, column=9, value=asset['act'])
                ws.cell(row=row_idx, column=10, value=asset['commission_date'])
                ws.cell(row=row_idx, column=11, value=asset['note'])

                if row_idx % 2 == 0:
                    for col in range(1, 12):
                        ws.cell(row=row_idx, column=col).fill = PatternFill(
                            start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')

            widths = [25, 15, 10, 12, 20, 20, 20, 12, 20, 15, 25]
            for col, width in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width

            ws.auto_filter.ref = f"A1:K{len(self.filtered_assets) + 1}"
            ws.freeze_panes = 'A2'
            wb.save(file_path)
            messagebox.showinfo("Успех", f"Отфильтрованные данные экспортированы в:\n{file_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось экспортировать: {str(e)}")

    # ================== АВТОСОХРАНЕНИЕ ==================
    def auto_save(self):
        self.save_to_excel()
        self.root.after(300000, self.auto_save)

    def on_closing(self):
        self.save_settings()
        if messagebox.askyesno("Выход", "Сохранить изменения перед выходом?"):
            if self.save_to_excel():
                self.root.destroy()
        else:
            self.root.destroy()

    # ================== ГОРЯЧИЕ КЛАВИШИ ==================
    def show_shortcuts_help(self):
        help_text = """Горячие клавиши:

Ctrl+N — добавить актив
Ctrl+E — редактировать выбранный актив
Ctrl+D — удалить выбранные активы
Ctrl+F — фокус на поле поиска
Ctrl+S — сохранить в Excel
Ctrl+P — печать отчёта

Двойной клик по группе — раскрыть/свернуть
Двойной клик по активу — редактировать
Одиночный клик по столбцу «Акт/Накладная» — открыть скан

Тема переключается кнопкой в правом нижнем углу."""
        messagebox.showinfo("Горячие клавиши", help_text)

    def setup_shortcuts(self):
        self.root.bind('<Control-n>', lambda e: self.add_asset())
        self.root.bind('<Control-e>', lambda e: self.edit_asset())
        self.root.bind('<Control-d>', lambda e: self.delete_asset())
        self.root.bind('<Control-f>', lambda e: self.focus_search())
        self.root.bind('<Control-s>', lambda e: self.save_to_excel())
        self.root.bind('<Control-p>', lambda e: self.print_report())

    def focus_search(self):
        self.search_entry.focus_set()


class AssetDialog:
    def __init__(self, parent, title, asset=None):
        self.dialog = tk.Toplevel(parent)
        self.dialog.title(title)
        self.dialog.geometry("550x700")
        self.dialog.configure(bg='#f0f0f0')
        self.dialog.resizable(False, False)

        self.result = None
        self.asset = asset or {}

        self.setup_ui()

        self.dialog.transient(parent)
        self.dialog.grab_set()
        self.dialog.focus_force()
        self.dialog.lift()

    def setup_ui(self):
        tk.Label(self.dialog, text="Информация об активе",
                font=('Segoe UI', 16, 'bold'),
                bg='#f0f0f0', fg='#333333').pack(pady=20)

        form_frame = tk.Frame(self.dialog, bg='#f0f0f0')
        form_frame.pack(fill=tk.BOTH, expand=True, padx=30)

        fields = [
            ('Наименование:', 'name', True),
            ('Инвентарный номер:', 'inventory', False),
            ('Количество:', 'quantity', False),
            ('Ед. измерения:', 'unit', False),
            ('Направление:', 'direction', True),
            ('Расположение:', 'location', True),
            ('Кому выдано:', 'issued_to', False),
            ('Статус:', 'status', False),
            ('Акт/Накладная:', 'act', False),
            ('Дата ввода:', 'commission_date', False),
            ('Примечание:', 'note', False)
        ]

        self.entries = {}

        for i, (label, key, required) in enumerate(fields):
            tk.Label(form_frame, text=label,
                    bg='#f0f0f0', font=('Segoe UI', 10)).grid(row=i, column=0, sticky='w', pady=5)

            if key == 'status':
                var = tk.StringVar(value=self.asset.get(key, 'Активен'))
                entry = ttk.Combobox(form_frame, textvariable=var,
                                    values=['Активен', 'В ремонте', 'Списан', 'Утеряно',
                                            'Требует ремонта', 'В резерве', 'В эксплуатации'],
                                    state='readonly', width=25)
            else:
                default = self.asset.get(key, '')
                if key == 'quantity':
                    default = self.asset.get(key, 1)
                elif key == 'unit':
                    default = self.asset.get(key, 'шт.')
                var = tk.StringVar(value=default)
                entry = tk.Entry(form_frame, textvariable=var, width=27,
                               font=('Segoe UI', 10))

            entry.grid(row=i, column=1, sticky='w', pady=5, padx=(10, 0))
            self.entries[key] = var

            if required:
                tk.Label(form_frame, text="*", fg='red',
                        bg='#f0f0f0', font=('Segoe UI', 10, 'bold')).grid(row=i, column=2, sticky='w')

        button_frame = tk.Frame(self.dialog, bg='#f0f0f0')
        button_frame.pack(pady=20)

        tk.Button(button_frame, text="💾 Сохранить",
                 command=self.save,
                 bg='#4CAF50', fg='white',
                 font=('Segoe UI', 10, 'bold'),
                 relief=tk.FLAT, cursor='hand2',
                 padx=20, pady=10).pack(side=tk.LEFT, padx=10)

        tk.Button(button_frame, text="Отмена",
                 command=self.dialog.destroy,
                 bg='#9E9E9E', fg='white',
                 font=('Segoe UI', 10),
                 relief=tk.FLAT, cursor='hand2',
                 padx=20, pady=10).pack(side=tk.LEFT, padx=10)

    def save(self):
        required_fields = ['name', 'direction', 'location']
        for field in required_fields:
            if not self.entries[field].get().strip():
                messagebox.showwarning("Внимание",
                                      f"Поле '{field}' обязательно для заполнения")
                return

        try:
            quantity = int(self.entries['quantity'].get().strip()) if self.entries['quantity'].get().strip() else 1
        except ValueError:
            messagebox.showerror("Ошибка", "Количество должно быть числом")
            return

        self.result = {
            'name': self.entries['name'].get().strip(),
            'inventory': self.entries['inventory'].get().strip(),
            'quantity': quantity,
            'unit': self.entries['unit'].get().strip() or 'шт.',
            'direction': self.entries['direction'].get().strip(),
            'location': self.entries['location'].get().strip(),
            'issued_to': self.entries['issued_to'].get().strip(),
            'status': self.entries['status'].get(),
            'act': self.entries['act'].get().strip(),
            'commission_date': self.entries['commission_date'].get().strip(),
            'note': self.entries['note'].get().strip()
        }

        self.dialog.destroy()


class MoveDialog:
    def __init__(self, parent, asset, manager):
        self.manager = manager
        self.asset = asset
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Перемещение актива")
        self.dialog.geometry("400x300")
        self.dialog.grab_set()

        tk.Label(self.dialog, text=f"Актив: {asset['name']} ({asset['inventory']})").pack(pady=10)

        tk.Label(self.dialog, text="Откуда:").pack()
        self.from_var = tk.StringVar(value=asset['location'])
        tk.Entry(self.dialog, textvariable=self.from_var).pack(pady=5)

        tk.Label(self.dialog, text="Куда:").pack()
        self.to_var = tk.StringVar()
        tk.Entry(self.dialog, textvariable=self.to_var).pack(pady=5)

        tk.Label(self.dialog, text="Дата (ГГГГ-ММ-ДД):").pack()
        self.date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        tk.Entry(self.dialog, textvariable=self.date_var).pack(pady=5)

        tk.Label(self.dialog, text="Примечание:").pack()
        self.note_var = tk.StringVar()
        tk.Entry(self.dialog, textvariable=self.note_var).pack(pady=5)

        tk.Button(self.dialog, text="Сохранить", command=self.save).pack(pady=20)

    def save(self):
        self.manager.record_move(
            self.asset['inventory'],
            self.from_var.get(),
            self.to_var.get(),
            self.date_var.get(),
            self.manager.current_user,
            self.note_var.get()
        )
        self.asset['location'] = self.to_var.get()
        self.manager.save_to_excel()
        self.manager.refresh_after_change()
        self.dialog.destroy()
        messagebox.showinfo("Успех", "Перемещение записано")


def main():
    root = tk.Tk()
    app = AssetManager(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()


if __name__ == "__main__":
    main()
